# @author: Tuana Cetinkaya
# @date: 09/02/2021
# this file handles excel operations
# parameters and return types specified for each method in case we need to switch to another file type in the future

# excel read and write libraries
# IMPORTANT VERSION CHECK! >> xlrd==1.2.0

# from openpyxl import Workbook
import openpyxl


# Report File is formatted as below:
# Code| Brand | Name | Price | Availability
class ExcelParser:
    data_file_name = ""
    report_file_name = ""
    report_sheet_name = ""
    workbook = None
    sheet = None

    def __init__(self, data_file_name, report_file_name, report_sheet_name):
        self.data_file_name = data_file_name
        self.report_file_name = report_file_name
        self.report_sheet_name = report_sheet_name

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    # @param file_path: excel file path with A column containing product links
    # @return a list of links retained from the file
    def get_links(self, website, data_sheet_name):
        links = []
        try:
            # alternative
            # df = pd.read_excel("myFile.xlsx", sheet_name=2, engine='openpyxl')

            book = openpyxl.load_workbook(self.data_file_name)
            sheet = book.active
            for row_idx in range(sheet.max_row):
                links.append(website + sheet.cell(row=row_idx + 1, column=1).value)

        except Exception as e:
            print(e)
            print("An Error Occurred While Reading the Sheet: "
                  "\n\tData File Containing The Product Links is Missing From the Directory"
                  "\n\tMake Sure Your File is in Excel Format and Located in the Same Directory as Our Program")
            exit(-1)

        return links

    def create_report_file(self):
        self.sheet['A1'] = "Product Code"
        self.sheet['B1'] = "Product Brand"
        self.sheet['C1'] = "Product Name"
        self.sheet['D1'] = "Product Price"
        self.sheet['E1'] = "Availability Percentage"

    def save_to_file(self, row, code, brand, name, price, availability):
        self.sheet['A' + str(row)] = code.text.strip()
        self.sheet['B' + str(row)] = brand.text.strip()
        self.sheet['C' + str(row)] = name.text.strip()
        self.sheet['D' + str(row)] = price.text.strip()
        self.sheet['E' + str(row)] = availability

        # to check the process uncomment the line below
        # print(row)

    def close_report_file(self):
        self.workbook.save(self.report_file_name)
